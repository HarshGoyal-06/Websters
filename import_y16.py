import psycopg2
from openpyxl import load_workbook
hostname='localhost'
database='STGI_Hackathon'
username='postgres'
pwd='Harsh@123'
port_id=5432
conn = None
cur = None
excel_file_path = "C:\\Users\\Dell\\OneDrive\\Desktop\\STGI\\H-1B_Disclosure_Data_FY16.xlsx"
batch_size = 1000
try:
    conn = psycopg2.connect(host = hostname, dbname = database, user = username, password = pwd, port = port_id)
    curr = conn.cursor()

    curr.execute('DROP TABLE IF EXISTS FY16_H1B')
    create_script='''CREATE TABLE FY16_H1B(
    CASE_NUMBER varchar(255),
    CASE_STATUS varchar(255),
    CASE_SUBMITTED DATE,
    DECISION_DATE DATE,
    VISA_CLASS varchar(255),
    EMPLOYMENT_START_DATE DATE,
    EMPLOYMENT_END_DATE DATE,
    EMPLOYER_NAME VARCHAR(255),
    EMPLOYER_ADDRESS VARCHAR(255),	
    EMPLOYER_CITY VARCHAR(255),
    EMPLOYER_STATE VARCHAR(255),
    EMPLOYER_POSTAL_CODE VARCHAR(255),
    EMPLOYER_COUNTRY varchar(255),
    EMPLOYER_PROVINCE VARCHAR(255),
    EMPLOYER_PHONE VARCHAR(255),
    EMPLOYER_PHONE_EXT VARCHAR(255),
    AGENT_ATTORNEY_NAME varchar(255),
    AGENT_ATTORNEY_CITY varchar(255),
    AGENT_ATTORNEY_STATE varchar(255),
    JOB_TITLE varchar(255),
    SOC_CODE varchar(255),
    SOC_NAME varchar(255),
    NAIC_CODE varchar(255),
    TOTAL_WORKERS varchar(255),
    FULL_TIME_POSITION varchar(255),
    PREVAILING_WAGE varchar(255),
    PW_UNIT_OF_PAY varchar(255),
    PW_WAGE_SOURCE varchar(255),
    PW_SOURCE_YEAR varchar(255),
    PW_SOURCE_OTHER varchar(255),
    WAGE_RATE_OF_PAY_FROM VARCHAR(255),
    WAGE_RATE_OF_PAY_TO varchar(255),
    WAGE_UNIT_OF_PAY varchar(255),
    H_1B_DEPENDENT varchar(255),
    WILLFUL_VIOLATOR varchar(255),
    WORKSITE_CITY varchar(255),
    WORKSITE_COUNTY varchar(255),
    WORKSITE_STATE varchar(255),
    WORKSITE_POSTAL_CODE varchar(255),
    ORIGINAL_CERT_DATE date)'''
    curr.execute(create_script)
    insert_script= 'INSERT INTO FY16_H1B(CASE_NUMBER,CASE_STATUS,CASE_SUBMITTED,DECISION_DATE,VISA_CLASS,EMPLOYMENT_START_DATE,EMPLOYMENT_END_DATE,EMPLOYER_COUNTRY,AGENT_ATTORNEY_NAME,AGENT_ATTORNEY_CITY,JOB_TITLE,SOC_CODE,SOC_NAME,NAIC_CODE,TOTAL_WORKERS,FULL_TIME_POSITION,PREVAILING_WAGE,PW_UNIT_OF_PAY,PW_WAGE_SOURCE,PW_SOURCE_YEAR,PW_SOURCE_OTHER,WAGE_RATE_OF_PAY_FROM,WAGE_RATE_OF_PAY_TO,WAGE_UNIT_OF_PAY,H_1B_DEPENDENT,WILLFUL_VIOLATOR,WORKSITE_COUNTY,ORIGINAL_CERT_DATE) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row = 2, values_only = True):
        data.append(row)
        if len(data) >= batch_size:
            curr.executemany(insert_script, data)
            conn.commit()
            data = []

    if data:
        curr.executemany(insert_script, data)
        conn.commit()

except Exception as error:
    print(error)
finally:
    if conn is not None:
        conn.close()
    if curr is not None:
        curr.close()